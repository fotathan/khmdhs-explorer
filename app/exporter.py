"""
exporter.py — turn selected tables into .xlsx output.

Two modes:
  - "workbook": one .xlsx, one sheet per table
  - "separate": one .xlsx per source file (tables from the same file share a
                workbook), bundled into a .zip
"""

from __future__ import annotations

import io
import os
import re
import zipfile
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _sheet_name(base: str, used: set[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub(" ", base).strip() or "Table"
    name = name[:31]
    candidate, i = name, 2
    while candidate.lower() in used:
        suffix = f" ({i})"
        candidate = name[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate.lower())
    return candidate


def _write_table(ws, table: dict) -> None:
    ws.append([f"{table['source']} — {table['locator']}"])
    ws["A1"].font = Font(italic=True, color="666666")
    for r_idx, row in enumerate(table["rows"], start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # style first data row as header
    for c_idx in range(1, table["n_cols"] + 1):
        cell = ws.cell(row=2, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # rough auto-width
    for c_idx in range(1, table["n_cols"] + 1):
        width = max(
            (len(table["rows"][r][c_idx - 1]) for r in range(table["n_rows"])),
            default=8,
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(width + 2, 9), 60)
    ws.freeze_panes = "A3"


def _workbook_for(tables: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for t in tables:
        base = os.path.splitext(os.path.basename(t["source"].split(" → ")[-1]))[0]
        ws = wb.create_sheet(_sheet_name(f"{base} {t['locator']}", used))
        _write_table(ws, t)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_filename(name: str) -> str:
    name = re.sub(r"[^\w\-. ()\u0370-\u03FF\u1F00-\u1FFF]", "_", name)
    return name.strip() or "tables"


def export_workbook(tables: list[dict]) -> tuple[bytes, str, str]:
    """All selected tables in one workbook."""
    data = _workbook_for(tables)
    return data, "tender-tables.xlsx", \
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_separate(tables: list[dict]) -> tuple[bytes, str, str]:
    """One workbook per source file, zipped."""
    by_source: dict[str, list[dict]] = defaultdict(list)
    for t in tables:
        by_source[t["source"]].append(t)

    buf = io.BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for source, group in by_source.items():
            base = os.path.splitext(os.path.basename(source.split(" → ")[-1]))[0]
            fname = _safe_filename(f"{base} - tables.xlsx")
            n = 2
            while fname.lower() in used_names:
                fname = _safe_filename(f"{base} - tables ({n}).xlsx")
                n += 1
            used_names.add(fname.lower())
            zf.writestr(fname, _workbook_for(group))
    return buf.getvalue(), "tender-tables.zip", "application/zip"
